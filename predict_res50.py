# -*- coding: utf-8 -*-
# Inception image recognition attempt
import tensorflow as tf
import numpy as np
import re
import os
import time
import sys
import math
import json
import logging

import glob
import openpyxl
import imagenet_preprocessing as ipp

from video_capture import VideoCaptureYUV

# Settings for the code
start_ID = 1
End_ID   = 50001
ORG = False
QF = 30
IS_HEVC = False

PATH_TO_SLIM = 'slim/'
PATH_TO_INCEPTION_MODEL = '/Users/hossam.amer/7aS7aS_Works/work/workspace/jpeg_ml_IPPP_real/bin/Build/Products/Release/machine_learning/visualize/visualize_inception_featureMaps/inception_model/'
PATH_TO_LABEL_LOOK_UP   = os.path.join(PATH_TO_INCEPTION_MODEL, 'imagenet_2012_challenge_label_map_proto.pbtxt')
PATH_TO_UID_LOOK_UP     = os.path.join(PATH_TO_INCEPTION_MODEL, 'imagenet_synset_to_human_label_map.txt')

# YUV Path
PATH_TO_RECONS = '/Volumes/MULTICOMHD2/set_yuv/Seq-RECONS/'

# JPEG Path
path_to_valid_images    = '/Volumes/MULTICOMHD2/validation_original/';
path_to_valid_QF_images = '/Volumes/MULTICOMHD2/validation_generated_QF/';

sys.path.append(PATH_TO_SLIM); 

from preprocessing import inception_preprocessing, vgg_preprocessing
from datasets import imagenet


os.environ['CUDA_VISIBLE_DEVICES'] = '-1'

# Change class ID to human readable label
class NodeLookup(object):
    def __init__(self):
        # Path for imagenet_2012_challenge_label_map_proto.pbtxt and imagenet_synset_to_human_label_map.txt
        label_lookup_path = PATH_TO_LABEL_LOOK_UP
        uid_lookup_path = PATH_TO_UID_LOOK_UP
        self.node_lookup = self.load(label_lookup_path, uid_lookup_path)

    def load(self, label_lookup_path, uid_lookup_path):
        # Load UID look up file
        proto_as_ascii_lines = tf.gfile.GFile(uid_lookup_path).readlines()
        uid_to_human = {}
        # Read Data line by line
        for line in proto_as_ascii_lines :
            # Remove \n
            line=line.strip('\n')
            # Split by \t
            parsed_items = line.split('\t')
            # Get the label number
            uid = parsed_items[0]
            # Get the human readable label name
            human_string = parsed_items[1]
            # Save the label
            uid_to_human[uid] = human_string

        # Load label look up file
        proto_as_ascii = tf.gfile.GFile(label_lookup_path).readlines()
        node_id_to_uid = {}
        for line in proto_as_ascii:
            if line.startswith('  target_class:'):
                # Get number 1-1000
                target_class = int(line.split(': ')[1])
            if line.startswith('  target_class_string:'):
                # Get UID n********
                target_class_string = line.split(': ')[1]
                # Save the relationship between number (1-1000) and the uid (n********)
                node_id_to_uid[target_class] = target_class_string[1:-2]

        # Build the relationship between number (1-1000) and the label
        node_id_to_name = {}
        for key, val in node_id_to_uid.items():
            # Get the label
            name = uid_to_human[val]
            # Build the relationship between number (1-1000) and the label
            node_id_to_name[key] = name
        return node_id_to_name

    # Input 1-1000 Return corresponding label
    def id_to_string(self, node_id):
        if node_id not in self.node_lookup:
            return ''
        return self.node_lookup[node_id]

def get_labels(json_file):
  """Get the set of possible labels for classification."""
  with open(json_file, "r") as labels_file:
    labels = json.load(labels_file)

  return labels

# Read pre-trained Inception-V3 to create graph
def create_graph():
  # the class that's been created from the textual definition in graph.proto
  # classify_image_graph_def.pb
  with tf.gfile.FastGFile('frozen_resnet_50_v2.pb', 'rb') as f:
        graph_def = tf.GraphDef()
        graph_def.ParseFromString(f.read())
        tf.import_graph_def(graph_def, name='')

def write_to_excel_sheet(gt, probs , current_human_gt_label, idx , sheet , col_idx = 4):
    predictions = np.squeeze(probs)
    N = -2000
    predictions = np.squeeze(predictions)
    top_5 = predictions.argsort()[N:][::-1]
    for rank, node_id in enumerate(top_5):
        human_string = gt[node_id]
        score = predictions[node_id]
        if(current_human_gt_label == human_string):
            current_rank =  1 + rank
            
            # Write only gt rank and gt probability
            sheet.cell(row=(idx-1)*6 + 2, column=4).value = current_rank
            sheet.cell(row=(idx-1)*6 + 2, column=5).value = score.item()

            print('%d: %s (rank = %d; score = %.5f)' % (ID, human_string, current_rank, score))
            break
    return current_rank    


def get_image_data(sess, imgID, QF, isCast = True):

    # Parse the YUV and convert it into RGB
    original_img_ID = imgID
    imgID = str(imgID).zfill(8)
    shard_num  = math.floor((original_img_ID - 1) / 10000)
    folder_num = math.ceil(original_img_ID/1000);
    if isCast:
        path_to_recons = PATH_TO_RECONS
        # Get files list to fetch the correct name
        filesList = glob.glob(path_to_recons + str(folder_num) + '/ILSVRC2012_val_' + imgID + '*.yuv')
        name = filesList[0].split('/')[-1]
        rgbStr = name.split('_')[5]
        width  = int(name.split('_')[-5])
        height = int(name.split('_')[-4])
        is_gray_str = name.split('_')[-3]
        
        image = path_to_recons + str(folder_num) + '/ILSVRC2012_val_' + imgID + '_' + str(width) + '_' + str(height) + '_' + rgbStr + '_' + str(QF) + '_1.yuv'
        size = (height, width) # height and then width
        videoObj = VideoCaptureYUV(image, size, isGrayScale=is_gray_str.__contains__('Y'))
        ret, yuv, rgb = videoObj.getYUVAndRGB()
        image_data = rgb
        image_data = tf.convert_to_tensor(image_data, dtype=tf.uint8)
    else:
        if QF == 110:
            image = path_to_valid_images + str(folder_num) + '/ILSVRC2012_val_' + imgID + '.JPEG'
        else:
            image = path_to_valid_QF_images + str(folder_num) + '/ILSVRC2012_val_' + imgID + '-QF-' + str(QF) + '.JPEG'            
        image_data = tf.read_file(image)
        image_data = tf.image.decode_jpeg(image_data, channels=3)
    
    image_data = inception_preprocessing.preprocess_image(image_data, 299, 299, is_training=False)
    image_data = tf.expand_dims(image_data, 0)
    
    return image_data

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3' # Hide the warning information from Tensorflow - annoying...
# Load Ground Truth JSON
# gt = get_labels("labellist.json")
gt = imagenet.create_readable_names_for_imagenet_labels()

############################ MAIN METHOD #######################
######

t = 0

path_to_excel = ''
if IS_HEVC:
    if ORG:
        path_to_excel = './Data/Accuracy Record ILSVRC2012 QF Res50 5-org.xlsm'
    else:
        path_to_excel = './Data/Accuracy Record ILSVRC2012 QF Res50 5-HEVC-' + str(QF) +'.xlsm'
else:
    if ORG:
        path_to_excel = './Data/Accuracy Record ILSVRC2012 QF Res50 5-org.xlsm'
    else:
        path_to_excel = './Data/Accuracy Record ILSVRC2012 QF Res50 5-' + str(QF) +'.xlsm'

# Start looping:
for ID in range(start_ID, End_ID):

    startTime = time.time()

    if (ID - 1) % 200 == 0:
        sess = tf.Session()
        create_graph()
        if ORG:
            wb = openpyxl.load_workbook(path_to_excel, read_only = False, keep_vba = True)
        else:
            wb = openpyxl.load_workbook(path_to_excel, read_only = False, keep_vba = True)
        sheet = wb['Sheet1']

    # Quality Factor Range
    for i in range(QF , QF-9, -10):
        
        # Get image data
        image_data = get_image_data(sess, ID, QF, IS_HEVC)

        # resnet 50: last layer is output as softmax
        softmax_tensor_res = sess.graph.get_tensor_by_name('resnet_v2_50/predictions/Reshape_1:0')

        # Input the image, obtain the softmax prob value（one shape=(1,1008) vector）
        predictions = sess.run(softmax_tensor_res,{'input:0': sess.run(image_data)})

        
        # Write to excelsheet
        current_gt_label = sheet["B"+str((ID-1)*6+2)].value
        write_to_excel_sheet(gt, predictions, current_gt_label, ID, sheet)
       
    t += time.time() - startTime
    if (ID) % 200 == 0:
        tf.reset_default_graph()
        sess.close()
        t = 0
        if ORG:
            wb.save(path_to_excel)
        else:
            wb.save(path_to_excel)

        print('Saving; Done with %d images in %f seconds..' % (ID, t))


